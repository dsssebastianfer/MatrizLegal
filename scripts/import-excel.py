"""
Importación desde Excel a SQLite (modo local) o Supabase (producción).

Uso:
  # Modo LOCAL (SQLite) — no requiere configuración:
  python scripts/import-excel.py

  # Modo SUPABASE — requiere credenciales:
  python scripts/import-excel.py --supabase
  # o con variables de entorno:
  SUPABASE_URL=https://xxx.supabase.co SUPABASE_SERVICE_ROLE_KEY=eyJ... python scripts/import-excel.py --supabase

Dependencias:
  pip install openpyxl requests
"""

import os
import sys
import sqlite3
import uuid
import json
from datetime import datetime, date

import openpyxl

# ── Configuración ──────────────────────────────────────────────────────────────
USE_SUPABASE = '--supabase' in sys.argv
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '03 MATRIZ LEGAL OCTUBRE 2025.xlsx')
SQLITE_PATH = os.path.join(os.path.dirname(__file__), '..', 'local.db')
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
# ──────────────────────────────────────────────────────────────────────────────

def to_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s if s else None

def map_estado(v) -> str:
    if v is None:
        return 'na'
    s = str(v).strip().lower()
    if s in ('1', 'cumple', 'si', 'sí'):
        return 'cumple'
    if s in ('parcial', '0.5'):
        return 'parcial'
    if s in ('0', 'no cumple', 'no'):
        return 'no_cumple'
    return 'na'

# ── Modo SQLite ────────────────────────────────────────────────────────────────

def init_sqlite(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS laws (
            id TEXT PRIMARY KEY,
            item INTEGER,
            codigo TEXT NOT NULL,
            titular TEXT,
            anio_publicacion TEXT,
            descripcion TEXT,
            mecanismo_evaluacion TEXT,
            fecha_ultima_evaluacion TEXT,
            estado_cumplimiento TEXT DEFAULT 'na',
            plan_accion TEXT,
            estado_plan_accion TEXT,
            periodicidad TEXT,
            vigencia TEXT,
            aplicacion TEXT,
            area TEXT,
            documento_nombre TEXT,
            documento_url TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS articles (
            id TEXT PRIMARY KEY,
            law_id TEXT NOT NULL REFERENCES laws(id) ON DELETE CASCADE,
            articulo TEXT,
            ambito_aplicacion TEXT,
            frecuencia_evaluacion TEXT,
            cumple INTEGER DEFAULT 0,
            parcial INTEGER DEFAULT 0,
            no_cumple INTEGER DEFAULT 0,
            na INTEGER DEFAULT 0,
            registro_evidencia TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id TEXT PRIMARY KEY,
            tabla TEXT NOT NULL,
            registro_id TEXT NOT NULL,
            accion TEXT NOT NULL,
            campo TEXT,
            valor_anterior TEXT,
            valor_nuevo TEXT,
            usuario_email TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()

def import_to_sqlite(wb):
    print(f'Conectando a SQLite: {SQLITE_PATH}')
    conn = sqlite3.connect(SQLITE_PATH)
    conn.execute('PRAGMA foreign_keys = ON')
    init_sqlite(conn)

    ws_resumen = wb['RESUMEN']
    rows = list(ws_resumen.iter_rows(min_row=3, values_only=True))
    all_sheets = set(wb.sheetnames)
    sheet_to_law_id = {}
    now = datetime.now().isoformat()

    print(f'Importando {len(rows)} filas del RESUMEN...')
    imported = 0
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        try:
            item = int(row[0]) if row[0] is not None else None
        except (ValueError, TypeError):
            continue
        codigo = to_str(row[1])
        if not codigo:
            continue

        law_id = str(uuid.uuid4())
        try:
            conn.execute("""
                INSERT INTO laws (id, item, codigo, titular, anio_publicacion, descripcion,
                    mecanismo_evaluacion, fecha_ultima_evaluacion, estado_cumplimiento,
                    plan_accion, estado_plan_accion, periodicidad, vigencia, aplicacion, area,
                    created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (law_id, item, codigo, to_str(row[2]), to_str(row[3]), to_str(row[4]),
                  to_str(row[5]), to_str(row[6]), map_estado(to_str(row[7])),
                  to_str(row[8]), to_str(row[9]), to_str(row[10]),
                  to_str(row[11]), to_str(row[12]), to_str(row[13]),
                  now, now))
            imported += 1
        except Exception as e:
            print(f'  ERROR ley {codigo}: {e}')
            continue

        print(f'  OK {item}: {codigo}')

        # Buscar hoja correspondiente
        for sheet in all_sheets:
            if sheet.lower() in ('resumen', 'control de cambios'):
                continue
            c = codigo.lower().replace(' ', '').replace('.', '').replace(',', '').replace('º', '').replace('n°', '').replace('nº', '')
            s = sheet.lower().replace(' ', '').replace('.', '').replace(',', '').replace('º', '').replace('n°', '').replace('nº', '')
            if c in s or s in c:
                sheet_to_law_id[sheet] = law_id
                break

    conn.commit()

    print(f'\nImportando artículos...')
    for sheet_name, law_id in sheet_to_law_id.items():
        try:
            ws = wb[sheet_name]
            count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                art_id = str(uuid.uuid4())
                conn.execute("""
                    INSERT INTO articles (id, law_id, articulo, ambito_aplicacion, frecuencia_evaluacion,
                        cumple, parcial, no_cumple, na, registro_evidencia, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (art_id, law_id,
                      to_str(row[0]) if len(row) > 0 else None,
                      to_str(row[1]) if len(row) > 1 else None,
                      to_str(row[2]) if len(row) > 2 else None,
                      1 if len(row) > 3 and row[3] not in (None, '', '--') else 0,
                      1 if len(row) > 4 and row[4] not in (None, '', '--') else 0,
                      1 if len(row) > 5 and row[5] not in (None, '', '--') else 0,
                      1 if len(row) > 6 and row[6] not in (None, '', '--') else 0,
                      to_str(row[7]) if len(row) > 7 else None,
                      now, now))
                count += 1
            conn.commit()
            print(f'  OK {sheet_name}: {count} articulos')
        except Exception as e:
            print(f'  ERROR en hoja {sheet_name}: {e}')

    conn.close()
    print(f'\nSQLite listo en: {SQLITE_PATH}')
    print(f'   Leyes importadas: {imported}')

# ── Modo Supabase ──────────────────────────────────────────────────────────────

def import_to_supabase(wb):
    import requests
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation',
    }
    def insert(table, data):
        r = requests.post(f'{SUPABASE_URL}/rest/v1/{table}', headers=headers, json=data)
        if r.status_code not in (200, 201):
            print(f'  ERROR {table}: {r.status_code} {r.text[:150]}')
            return None
        result = r.json()
        return result[0] if isinstance(result, list) else result

    ws_resumen = wb['RESUMEN']
    rows = list(ws_resumen.iter_rows(min_row=3, values_only=True))
    all_sheets = set(wb.sheetnames)
    sheet_to_law_id = {}

    print(f'Importando {len(rows)} filas a Supabase...')
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        try:
            item = int(row[0]) if row[0] is not None else None
        except (ValueError, TypeError):
            continue
        codigo = to_str(row[1])
        if not codigo:
            continue

        result = insert('laws', {
            'item': item, 'codigo': codigo, 'titular': to_str(row[2]),
            'anio_publicacion': to_str(row[3]), 'descripcion': to_str(row[4]),
            'mecanismo_evaluacion': to_str(row[5]), 'fecha_ultima_evaluacion': to_str(row[6]),
            'estado_cumplimiento': map_estado(to_str(row[7])),
            'plan_accion': to_str(row[8]), 'estado_plan_accion': to_str(row[9]),
            'periodicidad': to_str(row[10]), 'vigencia': to_str(row[11]),
            'aplicacion': to_str(row[12]), 'area': to_str(row[13]),
        })
        if not result:
            continue
        law_id = result['id']
        print(f'  OK {item}: {codigo}')

        for sheet in all_sheets:
            if sheet.lower() in ('resumen', 'control de cambios'):
                continue
            c = codigo.lower().replace(' ', '').replace('.', '').replace(',', '')
            s = sheet.lower().replace(' ', '').replace('.', '').replace(',', '')
            if c in s or s in c:
                sheet_to_law_id[sheet] = law_id
                break

    print('\nImportando artículos...')
    for sheet_name, law_id in sheet_to_law_id.items():
        try:
            ws = wb[sheet_name]
            count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                insert('articles', {
                    'law_id': law_id,
                    'articulo': to_str(row[0]) if len(row) > 0 else None,
                    'ambito_aplicacion': to_str(row[1]) if len(row) > 1 else None,
                    'frecuencia_evaluacion': to_str(row[2]) if len(row) > 2 else None,
                    'cumple': len(row) > 3 and row[3] not in (None, '', '--'),
                    'parcial': len(row) > 4 and row[4] not in (None, '', '--'),
                    'no_cumple': len(row) > 5 and row[5] not in (None, '', '--'),
                    'na': len(row) > 6 and row[6] not in (None, '', '--'),
                    'registro_evidencia': to_str(row[7]) if len(row) > 7 else None,
                })
                count += 1
            print(f'  OK {sheet_name}: {count} articulos')
        except Exception as e:
            print(f'  ERROR en {sheet_name}: {e}')

    print('\n✅ Supabase actualizado.')

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f'Leyendo {EXCEL_PATH} ...')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if USE_SUPABASE:
        print('Modo: SUPABASE\n')
        import_to_supabase(wb)
    else:
        print('Modo: SQLite local\n')
        import_to_sqlite(wb)

if __name__ == '__main__':
    main()
